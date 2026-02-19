#!/usr/bin/env python3
"""
Analisis de brecha 2025 vs Feb-2026 para Disponibilidad y UEBD.

Objetivo:
- Tomar un baseline (promedio diario) del 2025.
- Comparar contra Febrero 2026 (promedio diario real).
- Atribuir la brecha de UEBD y Disponibilidad a codigos (delta de horas/dia).
- Forzar que la suma de impactos atribuidos por codigo sea igual a la brecha (pp).
- Separar perdidas por:
  1) Disponibilidad
  2) UEBD (por codigo)
  3) Rendimiento F09 (filtrado en DrillPlan por "F09")
  4) Malla (valor manual, cuando se disponga)

Entradas:
- Archivo eventos 2025: DispUEBD_AllRigs_010125-0000_031225-2359(.csv)
- Archivo eventos 2026: DispUEBD_AllRigs_010126-0000_170226-2100(.csv)
- Opcional: mensual 2026 para leer UEBD/Disponibilidad objetivo de febrero.
"""

from __future__ import annotations

import argparse
import csv
import unicodedata
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from comparador_mensual_perforadoras import (
    normalize_equipo as normalize_equipo_mensual,
    parse_mensual_records,
    read_table as read_mensual_a1_table,
    recalculate_totals,
    resolve_table_file as resolve_mensual_a1_file,
)

DEFAULT_2025_FILE = "DispUEBD_AllRigs_010125-0000_031225-2359"
DEFAULT_2026_FILE = "DispUEBD_AllRigs_010126-0000_170226-2100"

MONTH_NAME_BY_NUM = {
    1: "Enero",
    2: "Febrero",
    3: "Marzo",
    4: "Abril",
    5: "Mayo",
    6: "Junio",
    7: "Julio",
    8: "Agosto",
    9: "Septiembre",
    10: "Octubre",
    11: "Noviembre",
    12: "Diciembre",
}


def normalize_text(value: Optional[str]) -> str:
    text = (value or "").strip().lower()
    decomposed = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in decomposed if not unicodedata.combining(ch))


def parse_datetime(value: Optional[str]) -> Optional[datetime]:
    if not value:
        return None
    raw = value.strip()
    if not raw:
        return None

    formats = (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M",
    )
    for fmt in formats:
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    return None


def parse_date(value: Optional[str]) -> Optional[date]:
    if not value:
        return None
    raw = value.strip()
    if not raw:
        return None
    try:
        return datetime.strptime(raw, "%Y-%m-%d").date()
    except ValueError:
        pass
    dt = parse_datetime(raw)
    if dt is not None:
        return dt.date()
    return None


def to_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    raw = str(value).strip()
    if not raw:
        return None
    raw = raw.replace("%", "").replace(",", ".")
    try:
        return float(raw)
    except ValueError:
        return None


def parse_ratio_value(value: Optional[float]) -> Optional[float]:
    if value is None:
        return None
    if value > 1:
        return value / 100.0
    if value < 0:
        return None
    return value


def resolve_input_csv_path(input_path: Path) -> Path:
    if input_path.exists():
        return input_path

    candidates: List[Path] = []
    if input_path.suffix == "":
        candidates.append(input_path.with_suffix(".csv"))
        candidates.append(input_path.with_suffix(".CSV"))
    elif input_path.suffix.lower() == ".csv":
        candidates.append(input_path.with_suffix(".CSV"))

    for candidate in candidates:
        if candidate.exists():
            return candidate

    attempted = ", ".join(str(c) for c in candidates) if candidates else "sin sugerencias"
    raise FileNotFoundError(
        f"No existe el archivo de entrada: {input_path}. Probados automaticamente: {attempted}"
    )


def get_operational_day(row: Dict[str, str], start_dt: Optional[datetime]) -> Optional[date]:
    work_day_started = parse_date(row.get("WorkDayStarted"))
    if work_day_started is not None:
        return work_day_started
    if start_dt is None:
        return None
    return (start_dt - timedelta(hours=21)).date()


def classify_bucket(row: Dict[str, str]) -> str:
    short_code = normalize_text(row.get("ShortCode"))
    planned = normalize_text(row.get("PlannedCodeName"))
    only_code_name = normalize_text(row.get("OnlyCodeName"))

    if short_code == "efectivo" or only_code_name.startswith("efectivo_"):
        return "efectivo"
    if short_code == "reserva":
        return "reserva"
    if short_code == "mantencion":
        if planned == "programada":
            return "mant_programada"
        return "mant_no_programada"
    return "otras"


def build_code_label(row: Dict[str, str]) -> str:
    code_number = (row.get("OnlyCodeNumber") or "").strip()
    code_name = (row.get("OnlyCodeName") or "").strip()
    code_name_alt = (row.get("CodeName") or "").strip()
    delay_data = (row.get("DelayData") or "").strip()

    if code_number and code_name:
        return f"{code_number}_{code_name}"
    if code_name:
        return code_name
    if code_name_alt:
        return code_name_alt
    if delay_data:
        return delay_data
    return "SIN_CODIGO"


def base_day_metrics() -> Dict[str, float]:
    return {
        "horas_totales": 0.0,
        "horas_operativas": 0.0,
        "horas_efectivo": 0.0,
        "horas_reserva": 0.0,
        "horas_mant_programada": 0.0,
        "horas_mant_no_programada": 0.0,
        "horas_otras": 0.0,
    }


def aggregate_daily_data(
    csv_path: Path,
    filter_year: Optional[int] = None,
    filter_month: Optional[int] = None,
) -> Dict[str, Any]:
    daily_metrics: Dict[date, Dict[str, float]] = {}
    daily_code_uebd: Dict[date, Dict[str, float]] = defaultdict(lambda: defaultdict(float))
    daily_code_disp: Dict[date, Dict[str, float]] = defaultdict(lambda: defaultdict(float))
    daily_code_f09: Dict[date, Dict[str, float]] = defaultdict(lambda: defaultdict(float))

    rows_read = 0

    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f, delimiter=";")
        for row in reader:
            rows_read += 1
            start_dt = parse_datetime(row.get("Time"))
            end_dt = parse_datetime(row.get("EndTime"))
            op_day = get_operational_day(row, start_dt)
            if op_day is None:
                continue

            if filter_year is not None and op_day.year != filter_year:
                continue
            if filter_month is not None and op_day.month != filter_month:
                continue

            dur_sec = to_float(row.get("Duration"))
            if dur_sec is None and start_dt is not None and end_dt is not None:
                dur_sec = (end_dt - start_dt).total_seconds()
            if dur_sec is None:
                dur_sec = 0.0
            dur_hours = max(dur_sec, 0.0) / 3600.0
            if dur_hours <= 0:
                continue

            if op_day not in daily_metrics:
                daily_metrics[op_day] = base_day_metrics()

            bucket = classify_bucket(row)
            code = build_code_label(row)
            drill_plan = normalize_text(row.get("DrillPlan"))
            is_f09 = "f09" in drill_plan

            metrics = daily_metrics[op_day]
            metrics["horas_totales"] += dur_hours

            if bucket == "mant_programada":
                metrics["horas_mant_programada"] += dur_hours
                daily_code_disp[op_day][code] += dur_hours
            elif bucket == "mant_no_programada":
                metrics["horas_mant_no_programada"] += dur_hours
                daily_code_disp[op_day][code] += dur_hours
            elif bucket == "efectivo":
                metrics["horas_efectivo"] += dur_hours
                metrics["horas_operativas"] += dur_hours
            elif bucket == "reserva":
                metrics["horas_reserva"] += dur_hours
                metrics["horas_operativas"] += dur_hours
                daily_code_uebd[op_day][code] += dur_hours
                if is_f09:
                    daily_code_f09[op_day][code] += dur_hours
            else:
                metrics["horas_otras"] += dur_hours
                metrics["horas_operativas"] += dur_hours
                daily_code_uebd[op_day][code] += dur_hours
                if is_f09:
                    daily_code_f09[op_day][code] += dur_hours

    days = sorted(daily_metrics.keys())
    day_count = len(days)
    totals = base_day_metrics()
    for d in days:
        for k, v in daily_metrics[d].items():
            totals[k] += v

    avg_hpd = {k: (totals[k] / day_count if day_count > 0 else 0.0) for k in totals}
    uebd_ratio = totals["horas_efectivo"] / totals["horas_operativas"] if totals["horas_operativas"] > 0 else 0.0
    disp_ratio = totals["horas_operativas"] / totals["horas_totales"] if totals["horas_totales"] > 0 else 0.0

    def avg_code_hours_per_day(daily_code_map: Dict[date, Dict[str, float]]) -> Dict[str, float]:
        accum: Dict[str, float] = defaultdict(float)
        for day in days:
            for code, hours in daily_code_map.get(day, {}).items():
                accum[code] += hours
        if day_count <= 0:
            return {}
        return {code: hours / day_count for code, hours in accum.items()}

    daily_rows = []
    for d in days:
        m = daily_metrics[d]
        daily_rows.append(
            {
                "fecha_operativa": d.isoformat(),
                **m,
                "uebd_ratio": m["horas_efectivo"] / m["horas_operativas"] if m["horas_operativas"] > 0 else 0.0,
                "uebd_pct": (m["horas_efectivo"] / m["horas_operativas"] * 100.0)
                if m["horas_operativas"] > 0
                else 0.0,
                "disponibilidad_ratio": m["horas_operativas"] / m["horas_totales"] if m["horas_totales"] > 0 else 0.0,
                "disponibilidad_pct": (m["horas_operativas"] / m["horas_totales"] * 100.0)
                if m["horas_totales"] > 0
                else 0.0,
            }
        )

    return {
        "rows_read": rows_read,
        "days": days,
        "day_count": day_count,
        "totals": totals,
        "avg_hpd": avg_hpd,
        "uebd_ratio": uebd_ratio,
        "disp_ratio": disp_ratio,
        "daily_rows": daily_rows,
        "uebd_code_hpd": avg_code_hours_per_day(daily_code_uebd),
        "disp_code_hpd": avg_code_hours_per_day(daily_code_disp),
        "f09_code_hpd": avg_code_hours_per_day(daily_code_f09),
    }


def aggregate_monthly_from_daily_rows(
    daily_rows: Iterable[Dict[str, Any]],
    target_year: int,
) -> List[Dict[str, Any]]:
    grouped: Dict[int, Dict[str, float]] = {}
    for row in daily_rows:
        d = parse_date(row.get("fecha_operativa"))
        if d is None or d.year != target_year:
            continue
        m = d.month
        if m not in grouped:
            grouped[m] = {
                "dias": 0.0,
                "horas_totales": 0.0,
                "horas_operativas": 0.0,
                "horas_efectivo": 0.0,
                "horas_reserva": 0.0,
                "horas_mant_programada": 0.0,
                "horas_mant_no_programada": 0.0,
                "horas_otras": 0.0,
            }
        rec = grouped[m]
        rec["dias"] += 1.0
        rec["horas_totales"] += float(row.get("horas_totales", 0.0))
        rec["horas_operativas"] += float(row.get("horas_operativas", 0.0))
        rec["horas_efectivo"] += float(row.get("horas_efectivo", 0.0))
        rec["horas_reserva"] += float(row.get("horas_reserva", 0.0))
        rec["horas_mant_programada"] += float(row.get("horas_mant_programada", 0.0))
        rec["horas_mant_no_programada"] += float(row.get("horas_mant_no_programada", 0.0))
        rec["horas_otras"] += float(row.get("horas_otras", 0.0))

    rows: List[Dict[str, Any]] = []
    for m in sorted(grouped):
        rec = grouped[m]
        ht = rec["horas_totales"]
        hop = rec["horas_operativas"]
        he = rec["horas_efectivo"]
        uebd_ratio = he / hop if hop > 0 else 0.0
        disp_ratio = hop / ht if ht > 0 else 0.0
        util_ratio = he / ht if ht > 0 else 0.0
        rows.append(
            {
                "anio": target_year,
                "mes_num": m,
                "mes": MONTH_NAME_BY_NUM.get(m, str(m)),
                "dias_con_datos": int(rec["dias"]),
                "horas_totales": ht,
                "horas_operativas": hop,
                "horas_efectivo": he,
                "horas_reserva": rec["horas_reserva"],
                "horas_mant_programada": rec["horas_mant_programada"],
                "horas_mant_no_programada": rec["horas_mant_no_programada"],
                "horas_otras": rec["horas_otras"],
                "disponibilidad_ratio_real": disp_ratio,
                "disponibilidad_pct_real": disp_ratio * 100.0,
                "utilizacion_ratio_real": util_ratio,
                "utilizacion_pct_real": util_ratio * 100.0,
                "uebd_ratio_real": uebd_ratio,
                "uebd_pct_real": uebd_ratio * 100.0,
            }
        )
    return rows


def read_table_any(path: Path, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            sample = f.read(4096)
            f.seek(0)
            delimiter = ";" if sample.count(";") >= sample.count(",") else ","
            reader = csv.DictReader(f, delimiter=delimiter)
            return [dict(row) for row in reader]

    if path.suffix.lower() in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        try:
            from openpyxl import load_workbook
        except Exception as exc:  # pragma: no cover
            raise RuntimeError(f"No se pudo importar openpyxl: {exc}") from exc

        wb = load_workbook(path, data_only=True)
        ws = wb[sheet_name] if (sheet_name and sheet_name in wb.sheetnames) else wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        header_row_idx = 0
        for idx, row in enumerate(rows):
            if row and any(cell is not None and str(cell).strip() for cell in row):
                header_row_idx = idx
                break
        headers = [str(c).strip() if c is not None else "" for c in rows[header_row_idx]]
        table = []
        for row in rows[header_row_idx + 1 :]:
            if row is None or not any(cell is not None and str(cell).strip() for cell in row):
                continue
            record = {}
            for i, h in enumerate(headers):
                if not h:
                    continue
                record[h] = row[i] if i < len(row) else None
            table.append(record)
        return table

    raise ValueError(f"Formato no soportado para archivo mensual: {path}")


def parse_month_value(value: Any) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        iv = int(value)
        if 1 <= iv <= 12:
            return iv
    txt = normalize_text(str(value))
    if txt.isdigit():
        iv = int(txt)
        if 1 <= iv <= 12:
            return iv
    month_map = {
        "enero": 1,
        "febrero": 2,
        "marzo": 3,
        "abril": 4,
        "mayo": 5,
        "junio": 6,
        "julio": 7,
        "agosto": 8,
        "septiembre": 9,
        "setiembre": 9,
        "octubre": 10,
        "noviembre": 11,
        "diciembre": 12,
    }
    return month_map.get(txt)


def find_col(headers: Iterable[str], candidates: Iterable[str]) -> Optional[str]:
    normalized_map = {h: normalize_text(h) for h in headers if h}
    for candidate in candidates:
        c = normalize_text(candidate)
        for h, nh in normalized_map.items():
            if c in nh:
                return h
    return None


def extract_monthly_targets(
    mensual_path: Path,
    year: int,
    month: int,
    sheet_name: Optional[str] = None,
    rig_preference: str = "TOTAL",
) -> Dict[str, Optional[float]]:
    table = read_table_any(mensual_path, sheet_name=sheet_name)
    if not table:
        return {"uebd_ratio": None, "disp_ratio": None}

    headers = list(table[0].keys())
    col_year = find_col(headers, ["anio", "year"])
    col_month = find_col(headers, ["mes", "month"])
    col_uebd = find_col(headers, ["uebd"])
    col_disp = find_col(headers, ["disponibilidad", "disp"])
    col_rig = find_col(headers, ["perforadora", "rig", "equipo", "flota", "total"])

    matches = []
    for row in table:
        row_year = to_float(row.get(col_year)) if col_year else None
        row_month = parse_month_value(row.get(col_month)) if col_month else None
        if row_year is None or row_month is None:
            continue
        if int(row_year) != int(year) or int(row_month) != int(month):
            continue
        matches.append(row)

    if not matches:
        return {"uebd_ratio": None, "disp_ratio": None}

    preferred_row = matches[0]
    if col_rig:
        preferred_norm = normalize_text(rig_preference)
        for row in matches:
            rig_val = normalize_text(row.get(col_rig))
            if rig_val == preferred_norm:
                preferred_row = row
                break
            if preferred_norm == "total" and rig_val in {"flota", "total"}:
                preferred_row = row
                break

    uebd_ratio = parse_ratio_value(to_float(preferred_row.get(col_uebd))) if col_uebd else None
    disp_ratio = parse_ratio_value(to_float(preferred_row.get(col_disp))) if col_disp else None
    return {"uebd_ratio": uebd_ratio, "disp_ratio": disp_ratio}


def build_monthly_targets_from_long_table(
    mensual_path: Path,
    year: int,
    sheet_name: Optional[str] = None,
    rig_preference: str = "TOTAL",
) -> List[Dict[str, Any]]:
    table = read_table_any(mensual_path, sheet_name=sheet_name)
    if not table:
        return []

    headers = list(table[0].keys())
    col_year = find_col(headers, ["anio", "year"])
    col_month = find_col(headers, ["mes", "month"])
    col_uebd = find_col(headers, ["uebd"])
    col_disp = find_col(headers, ["disponibilidad", "disp"])
    col_util = find_col(headers, ["utilizacion", "utilización", "util"])
    col_rig = find_col(headers, ["perforadora", "rig", "equipo", "flota", "total"])
    if not col_year or not col_month:
        return []

    preferred_norm = normalize_text(rig_preference)
    grouped: Dict[int, Dict[str, Any]] = {}
    for row in table:
        row_year = to_float(row.get(col_year))
        row_month = parse_month_value(row.get(col_month))
        if row_year is None or row_month is None:
            continue
        if int(row_year) != int(year):
            continue

        rig_val = normalize_text(row.get(col_rig)) if col_rig else ""
        is_preferred = (
            rig_val == preferred_norm
            or (preferred_norm == "total" and rig_val in {"total", "flota"})
            or (not col_rig)
        )
        if not is_preferred:
            continue

        uebd_ratio = parse_ratio_value(to_float(row.get(col_uebd))) if col_uebd else None
        disp_ratio = parse_ratio_value(to_float(row.get(col_disp))) if col_disp else None
        util_ratio = parse_ratio_value(to_float(row.get(col_util))) if col_util else None
        grouped[row_month] = {
            "anio": year,
            "mes_num": row_month,
            "mes": MONTH_NAME_BY_NUM.get(row_month, str(row_month)),
            "disponibilidad_ratio_obj": disp_ratio,
            "disponibilidad_pct_obj": (disp_ratio * 100.0) if disp_ratio is not None else None,
            "utilizacion_ratio_obj": util_ratio,
            "utilizacion_pct_obj": (util_ratio * 100.0) if util_ratio is not None else None,
            "uebd_ratio_obj": uebd_ratio,
            "uebd_pct_obj": (uebd_ratio * 100.0) if uebd_ratio is not None else None,
        }

    return [grouped[m] for m in sorted(grouped)]


def build_monthly_targets_from_mensual_a1(
    mensual_a1_path: Path,
    compare_year: int,
    sheet_name: Optional[str],
    excluded_equipos_csv: str,
) -> Dict[str, Any]:
    """Devuelve objetivos mensuales (disp/util/uebd) desde mensual A1 recalculado."""
    resolved_path = resolve_mensual_a1_file(mensual_a1_path)
    records = parse_mensual_records(
        read_mensual_a1_table(resolved_path, sheet_name),
        source_year=compare_year,
    )
    excluded = {
        normalize_equipo_mensual(x)
        for x in excluded_equipos_csv.split(",")
        if str(x).strip()
    }
    rec_rows, _cmp_rows = recalculate_totals(
        records,
        excluded_equipo_norms=excluded,
        source_year=compare_year,
    )

    by_month: Dict[int, Dict[str, Any]] = {}
    for row in rec_rows:
        month_num = int(row.get("mes_num", 0))
        if month_num not in by_month:
            by_month[month_num] = {
                "anio": compare_year,
                "mes_num": month_num,
                "mes": MONTH_NAME_BY_NUM.get(month_num, str(month_num)),
                "disponibilidad_ratio_obj": None,
                "disponibilidad_pct_obj": None,
                "utilizacion_ratio_obj": None,
                "utilizacion_pct_obj": None,
                "uebd_ratio_obj": None,
                "uebd_pct_obj": None,
            }

        idx = str(row.get("indice_key", "")).strip().lower()
        val = to_float(row.get("valor_total_recalculado_sin_excluidas"))
        if val is None:
            continue
        if idx == "disponibilidad":
            ratio = parse_ratio_value(val)
            by_month[month_num]["disponibilidad_ratio_obj"] = ratio
            by_month[month_num]["disponibilidad_pct_obj"] = (ratio * 100.0) if ratio is not None else None
        elif idx == "utilizacion":
            ratio = parse_ratio_value(val)
            by_month[month_num]["utilizacion_ratio_obj"] = ratio
            by_month[month_num]["utilizacion_pct_obj"] = (ratio * 100.0) if ratio is not None else None
        elif idx == "uebd":
            ratio = parse_ratio_value(val)
            by_month[month_num]["uebd_ratio_obj"] = ratio
            by_month[month_num]["uebd_pct_obj"] = (ratio * 100.0) if ratio is not None else None

    for month_num, rec in by_month.items():
        # Si UEBD no viene explicito, usar Utilizacion / Disponibilidad.
        if rec["uebd_ratio_obj"] is None:
            util = rec["utilizacion_ratio_obj"]
            disp = rec["disponibilidad_ratio_obj"]
            if util is not None and disp is not None and disp > 0:
                rec["uebd_ratio_obj"] = util / disp
                rec["uebd_pct_obj"] = rec["uebd_ratio_obj"] * 100.0

    rows = [by_month[m] for m in sorted(by_month)]
    return {"rows": rows, "source_path": str(resolved_path)}


def extract_targets_from_mensual_a1(
    mensual_a1_path: Path,
    compare_year: int,
    compare_month: int,
    sheet_name: Optional[str],
    excluded_equipos_csv: str,
) -> Dict[str, Optional[float]]:
    """
    Extrae objetivos desde mensual A1 (wide format) recalculando totales sin equipos excluidos.
    Devuelve disponibilidad y UEBD objetivo para el mes comparado.
    """
    monthly_targets = build_monthly_targets_from_mensual_a1(
        mensual_a1_path=mensual_a1_path,
        compare_year=compare_year,
        sheet_name=sheet_name,
        excluded_equipos_csv=excluded_equipos_csv,
    )
    target_rows = monthly_targets["rows"]
    target_row = next((r for r in target_rows if int(r.get("mes_num", 0)) == int(compare_month)), None)
    if target_row is None:
        return {
            "uebd_ratio": None,
            "disp_ratio": None,
            "util_ratio": None,
            "source_path": monthly_targets["source_path"],
            "rows": target_rows,
        }

    return {
        "uebd_ratio": target_row.get("uebd_ratio_obj"),
        "disp_ratio": target_row.get("disponibilidad_ratio_obj"),
        "util_ratio": target_row.get("utilizacion_ratio_obj"),
        "source_path": monthly_targets["source_path"],
        "rows": target_rows,
    }


def build_monthly_comparison_rows(
    monthly_real_rows: List[Dict[str, Any]],
    monthly_target_rows: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    real_by_month = {int(r["mes_num"]): r for r in monthly_real_rows}
    target_by_month = {int(r["mes_num"]): r for r in monthly_target_rows}
    months = sorted(set(real_by_month) | set(target_by_month))
    rows: List[Dict[str, Any]] = []
    for m in months:
        real = real_by_month.get(m, {})
        obj = target_by_month.get(m, {})
        uebd_obj = obj.get("uebd_ratio_obj")
        disp_obj = obj.get("disponibilidad_ratio_obj")
        util_obj = obj.get("utilizacion_ratio_obj")
        uebd_real = real.get("uebd_ratio_real")
        disp_real = real.get("disponibilidad_ratio_real")
        util_real = real.get("utilizacion_ratio_real")

        uebd_gap_pp = (
            (uebd_obj - uebd_real) * 100.0
            if (uebd_obj is not None and uebd_real is not None)
            else None
        )
        disp_gap_pp = (
            (disp_obj - disp_real) * 100.0
            if (disp_obj is not None and disp_real is not None)
            else None
        )
        util_gap_pp = (
            (util_obj - util_real) * 100.0
            if (util_obj is not None and util_real is not None)
            else None
        )

        hor_op = float(real.get("horas_operativas", 0.0) or 0.0)
        hor_tot = float(real.get("horas_totales", 0.0) or 0.0)
        perdida_uebd_h = (max(uebd_gap_pp, 0.0) / 100.0 * hor_op) if uebd_gap_pp is not None else None
        perdida_disp_h = (max(disp_gap_pp, 0.0) / 100.0 * hor_tot) if disp_gap_pp is not None else None

        rows.append(
            {
                "anio": int(real.get("anio", obj.get("anio", 0) or 0)),
                "mes_num": m,
                "mes": MONTH_NAME_BY_NUM.get(m, str(m)),
                "dias_reales_con_datos": int(real.get("dias_con_datos", 0) or 0),
                "horas_totales_real": float(real.get("horas_totales", 0.0) or 0.0),
                "horas_operativas_real": float(real.get("horas_operativas", 0.0) or 0.0),
                "horas_efectivo_real": float(real.get("horas_efectivo", 0.0) or 0.0),
                "disponibilidad_obj_ratio": disp_obj,
                "disponibilidad_obj_pct": (disp_obj * 100.0) if disp_obj is not None else None,
                "disponibilidad_real_ratio": disp_real,
                "disponibilidad_real_pct": (disp_real * 100.0) if disp_real is not None else None,
                "disponibilidad_gap_pp": disp_gap_pp,
                "utilizacion_obj_ratio": util_obj,
                "utilizacion_obj_pct": (util_obj * 100.0) if util_obj is not None else None,
                "utilizacion_real_ratio": util_real,
                "utilizacion_real_pct": (util_real * 100.0) if util_real is not None else None,
                "utilizacion_gap_pp": util_gap_pp,
                "uebd_obj_ratio": uebd_obj,
                "uebd_obj_pct": (uebd_obj * 100.0) if uebd_obj is not None else None,
                "uebd_real_ratio": uebd_real,
                "uebd_real_pct": (uebd_real * 100.0) if uebd_real is not None else None,
                "uebd_gap_pp": uebd_gap_pp,
                "perdida_horas_uebd_mes": perdida_uebd_h,
                "perdida_horas_disponibilidad_mes": perdida_disp_h,
            }
        )
    return rows


def build_code_delta_rows(
    baseline_hpd_by_code: Dict[str, float],
    compare_hpd_by_code: Dict[str, float],
) -> List[Dict[str, float]]:
    all_codes = sorted(set(baseline_hpd_by_code) | set(compare_hpd_by_code))
    rows: List[Dict[str, float]] = []
    for code in all_codes:
        b = float(baseline_hpd_by_code.get(code, 0.0))
        c = float(compare_hpd_by_code.get(code, 0.0))
        rows.append(
            {
                "codigo": code,
                "baseline_horas_dia": b,
                "comparado_horas_dia": c,
                "delta_horas_dia": c - b,
                "delta_horas_dia_positivo": max(c - b, 0.0),
            }
        )
    rows.sort(key=lambda r: r["delta_horas_dia"], reverse=True)
    return rows


def attribute_gap_to_codes(
    delta_rows: List[Dict[str, float]],
    denominator_hpd: float,
    gap_pp: float,
    compared_days: int,
) -> List[Dict[str, float]]:
    if denominator_hpd <= 0:
        for row in delta_rows:
            row["impacto_raw_pp"] = 0.0
            row["factor_escalamiento"] = 0.0
            row["impacto_atribuido_pp"] = 0.0
            row["perdida_horas_mes_atribuida"] = 0.0
        return delta_rows

    raw_sum_pp = 0.0
    for row in delta_rows:
        raw_pp = row["delta_horas_dia_positivo"] / denominator_hpd * 100.0
        row["impacto_raw_pp"] = raw_pp
        raw_sum_pp += raw_pp

    factor = (gap_pp / raw_sum_pp) if (raw_sum_pp > 0 and gap_pp > 0) else 0.0
    for row in delta_rows:
        impacto_pp = row["impacto_raw_pp"] * factor
        row["factor_escalamiento"] = factor
        row["impacto_atribuido_pp"] = impacto_pp
        perdida_hpd = impacto_pp / 100.0 * denominator_hpd
        row["perdida_horas_dia_atribuida"] = perdida_hpd
        row["perdida_horas_mes_atribuida"] = perdida_hpd * compared_days

    ordered = sorted(delta_rows, key=lambda r: r.get("impacto_atribuido_pp", 0.0), reverse=True)
    total_impact = sum(max(float(r.get("impacto_atribuido_pp", 0.0)), 0.0) for r in ordered)
    acum = 0.0
    for idx, row in enumerate(ordered, start=1):
        imp = max(float(row.get("impacto_atribuido_pp", 0.0)), 0.0)
        acum += imp
        row["ranking_impacto"] = idx
        row["participacion_gap_pct"] = (imp / total_impact * 100.0) if total_impact > 0 else 0.0
        row["impacto_acumulado_pp"] = acum
    return ordered


def sum_field(rows: Iterable[Dict[str, float]], field: str) -> float:
    return sum(float(r.get(field, 0.0)) for r in rows)


def format_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:.6f}"
    return str(value)


def write_csv(path: Path, rows: List[Dict[str, Any]], fieldnames: List[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=";")
        writer.writeheader()
        for row in rows:
            writer.writerow({k: format_value(row.get(k, "")) for k in fieldnames})


def top_contributions(rows: List[Dict[str, float]], top_n: int) -> List[Tuple[str, float]]:
    ordered = sorted(rows, key=lambda r: r.get("impacto_atribuido_pp", 0.0), reverse=True)
    selected = ordered[:top_n]
    other = ordered[top_n:]
    contribs = [(str(r["codigo"]), -float(r["impacto_atribuido_pp"])) for r in selected if r["impacto_atribuido_pp"] > 0]
    other_sum = sum(float(r.get("impacto_atribuido_pp", 0.0)) for r in other if r.get("impacto_atribuido_pp", 0.0) > 0)
    if other_sum > 0:
        contribs.append(("Otros", -other_sum))
    return contribs


def try_import_pyplot():
    try:
        import matplotlib

        matplotlib.use("Agg")
        import matplotlib.pyplot as plt

        return plt, ""
    except Exception as exc:  # pragma: no cover
        return None, str(exc)


def draw_gap_waterfall(
    output_path: Path,
    title: str,
    start_pct: float,
    final_pct: float,
    contributions: List[Tuple[str, float]],
) -> Tuple[bool, str]:
    plt, err = try_import_pyplot()
    if plt is None:
        return False, err

    labels = ["Objetivo"] + [c[0] for c in contributions] + ["Real"]
    fig_width = max(10.0, len(labels) * 0.9)
    fig, ax = plt.subplots(figsize=(fig_width, 6.0))

    ax.bar(0, start_pct, color="#2ca02c", edgecolor="black")
    ax.text(0, start_pct + 0.8, f"{start_pct:.2f}", ha="center", va="bottom", fontsize=8)

    running = start_pct
    for idx, (_label, delta_pp) in enumerate(contributions, start=1):
        nxt = running + delta_pp
        bottom = min(running, nxt)
        height = abs(delta_pp)
        color = "#d62728" if delta_pp < 0 else "#2ca02c"
        ax.bar(idx, height, bottom=bottom, color=color, edgecolor="black")
        ax.text(idx, bottom + height + 0.6, f"{delta_pp:.2f}", ha="center", va="bottom", fontsize=8)
        running = nxt

    final_idx = len(labels) - 1
    ax.bar(final_idx, final_pct, color="#1f77b4", edgecolor="black")
    ax.text(final_idx, final_pct + 0.8, f"{final_pct:.2f}", ha="center", va="bottom", fontsize=8)

    ax.set_xticks(range(len(labels)))
    ax.set_xticklabels(labels, rotation=35, ha="right")
    ax.set_ylabel("Puntos porcentuales (%)")
    ax.set_title(title)
    ax.grid(axis="y", linestyle="--", alpha=0.3)
    fig.tight_layout()

    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, dpi=150)
    plt.close(fig)
    return True, ""


def draw_monthly_obj_vs_real_lines(
    output_path: Path,
    rows: List[Dict[str, Any]],
    obj_key_pct: str,
    real_key_pct: str,
    title: str,
) -> Tuple[bool, str]:
    plt, err = try_import_pyplot()
    if plt is None:
        return False, err
    if not rows:
        return False, "Sin datos mensuales para comparar"

    xs = [str(r.get("mes", r.get("mes_num", ""))) for r in rows]
    obj_vals = [float(v) if v is not None else float("nan") for v in (r.get(obj_key_pct) for r in rows)]
    real_vals = [float(v) if v is not None else float("nan") for v in (r.get(real_key_pct) for r in rows)]

    x_idx = list(range(len(xs)))
    fig, ax = plt.subplots(figsize=(10, 5.5))
    ax.plot(x_idx, obj_vals, marker="o", linewidth=2, label="Objetivo mensual")
    ax.plot(x_idx, real_vals, marker="o", linewidth=2, label="Real eventos")
    ax.set_xticks(x_idx)
    ax.set_xticklabels(xs, rotation=25, ha="right")
    ax.set_ylabel("%")
    ax.set_title(title)
    ax.grid(alpha=0.3, linestyle="--")
    ax.legend()
    fig.tight_layout()

    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, dpi=150)
    plt.close(fig)
    return True, ""


def draw_monthly_gap_bars(
    output_path: Path,
    rows: List[Dict[str, Any]],
    gap_key_pp: str,
    title: str,
) -> Tuple[bool, str]:
    plt, err = try_import_pyplot()
    if plt is None:
        return False, err
    if not rows:
        return False, "Sin datos mensuales para gap"

    xs = [str(r.get("mes", r.get("mes_num", ""))) for r in rows]
    vals = [float(r.get(gap_key_pp) or 0.0) for r in rows]
    x_idx = list(range(len(xs)))

    fig, ax = plt.subplots(figsize=(10, 5.5))
    colors = ["#d62728" if v > 0 else "#2ca02c" for v in vals]
    ax.bar(x_idx, vals, color=colors, edgecolor="black")
    for i, v in enumerate(vals):
        ax.text(i, v + (0.2 if v >= 0 else -0.2), f"{v:.2f}", ha="center", va="bottom" if v >= 0 else "top", fontsize=8)
    ax.set_xticks(x_idx)
    ax.set_xticklabels(xs, rotation=25, ha="right")
    ax.set_ylabel("Gap (pp)")
    ax.set_title(title)
    ax.axhline(0, color="black", linewidth=1)
    ax.grid(axis="y", alpha=0.3, linestyle="--")
    fig.tight_layout()

    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, dpi=150)
    plt.close(fig)
    return True, ""


def draw_top_code_impact_bars(
    output_path: Path,
    rows: List[Dict[str, Any]],
    title: str,
    top_n: int = 15,
) -> Tuple[bool, str]:
    plt, err = try_import_pyplot()
    if plt is None:
        return False, err

    top_rows = [r for r in rows if float(r.get("impacto_atribuido_pp", 0.0)) > 0][:top_n]
    if not top_rows:
        return False, "Sin datos positivos para graficar"

    labels = [str(r.get("codigo", "")) for r in top_rows][::-1]
    vals = [float(r.get("impacto_atribuido_pp", 0.0)) for r in top_rows][::-1]

    fig, ax = plt.subplots(figsize=(11, max(5.0, len(labels) * 0.45)))
    bars = ax.barh(range(len(labels)), vals, color="#d62728", edgecolor="black")
    ax.set_yticks(range(len(labels)))
    ax.set_yticklabels(labels)
    ax.set_xlabel("Impacto atribuido (pp)")
    ax.set_title(title)
    ax.grid(axis="x", alpha=0.3, linestyle="--")
    for i, b in enumerate(bars):
        v = vals[i]
        ax.text(v + 0.05, b.get_y() + b.get_height() / 2, f"{v:.2f}", va="center", fontsize=8)
    fig.tight_layout()

    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, dpi=150)
    plt.close(fig)
    return True, ""


def try_import_xlsxwriter():
    try:
        import xlsxwriter

        return xlsxwriter, ""
    except Exception as exc:  # pragma: no cover
        return None, str(exc)


def write_excel(
    output_xlsx: Path,
    resumen_rows: List[Dict[str, Any]],
    uebd_rows: List[Dict[str, Any]],
    disp_rows: List[Dict[str, Any]],
    f09_rows: List[Dict[str, Any]],
    daily_2025_rows: List[Dict[str, Any]],
    daily_feb_rows: List[Dict[str, Any]],
    monthly_comparison_rows: List[Dict[str, Any]],
    perdidas_resumen_rows: List[Dict[str, Any]],
) -> Tuple[bool, str]:
    xlsxwriter, err = try_import_xlsxwriter()
    if xlsxwriter is None:
        return False, err

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb = xlsxwriter.Workbook(str(output_xlsx))
    header_fmt = wb.add_format({"bold": True, "bg_color": "#D9E1F2", "border": 1})
    num_fmt = wb.add_format({"num_format": "0.0000"})
    pct_fmt = wb.add_format({"num_format": "0.00"})

    def add_sheet(name: str, rows: List[Dict[str, Any]]):
        ws = wb.add_worksheet(name[:31])
        if not rows:
            ws.write(0, 0, "Sin datos")
            return ws, []
        cols = list(rows[0].keys())
        for c, col in enumerate(cols):
            ws.write(0, c, col, header_fmt)
        for r, row in enumerate(rows, start=1):
            for c, col in enumerate(cols):
                val = row.get(col, "")
                if isinstance(val, float):
                    fmt = pct_fmt if col.endswith("_pct") or col.endswith("_pp") else num_fmt
                    ws.write_number(r, c, val, fmt)
                elif isinstance(val, int):
                    ws.write_number(r, c, val)
                else:
                    ws.write(r, c, str(val))
        for c, col in enumerate(cols):
            ws.set_column(c, c, max(12, min(42, len(col) + 2)))
        return ws, cols

    ws_resumen, cols_resumen = add_sheet("Resumen_Brecha", resumen_rows)
    ws_cmp, cols_cmp = add_sheet("Comparacion_Mensual", monthly_comparison_rows)
    ws_perd, cols_perd = add_sheet("Perdidas_Resumen", perdidas_resumen_rows)

    top_uebd = [r for r in uebd_rows if float(r.get("impacto_atribuido_pp", 0.0)) > 0][:20]
    top_disp = [r for r in disp_rows if float(r.get("impacto_atribuido_pp", 0.0)) > 0][:20]
    ws_top_uebd, cols_top_uebd = add_sheet("Top_UEBD_Codigos", top_uebd)
    ws_top_disp, cols_top_disp = add_sheet("Top_Disp_Codigos", top_disp)

    add_sheet("Aporte_UEBD_Codigos", uebd_rows)
    add_sheet("Aporte_Disp_Codigos", disp_rows)
    add_sheet("Perdida_F09_Codigos", f09_rows)
    add_sheet("Diario_2025", daily_2025_rows)
    add_sheet("Diario_Feb2026", daily_feb_rows)

    if monthly_comparison_rows:
        last = len(monthly_comparison_rows)
        c_mes = cols_cmp.index("mes")
        c_uebd_obj = cols_cmp.index("uebd_obj_pct")
        c_uebd_real = cols_cmp.index("uebd_real_pct")
        c_uebd_gap = cols_cmp.index("uebd_gap_pp")
        c_disp_obj = cols_cmp.index("disponibilidad_obj_pct")
        c_disp_real = cols_cmp.index("disponibilidad_real_pct")

        chart_uebd = wb.add_chart({"type": "line"})
        chart_uebd.add_series(
            {
                "name": "UEBD Objetivo (%)",
                "categories": ["Comparacion_Mensual", 1, c_mes, last, c_mes],
                "values": ["Comparacion_Mensual", 1, c_uebd_obj, last, c_uebd_obj],
            }
        )
        chart_uebd.add_series(
            {
                "name": "UEBD Real (%)",
                "categories": ["Comparacion_Mensual", 1, c_mes, last, c_mes],
                "values": ["Comparacion_Mensual", 1, c_uebd_real, last, c_uebd_real],
            }
        )
        chart_uebd.set_title({"name": "UEBD mensual: Objetivo vs Real"})
        chart_uebd.set_y_axis({"name": "%"})
        chart_uebd.set_legend({"position": "bottom"})
        ws_cmp.insert_chart("AA2", chart_uebd, {"x_scale": 1.15, "y_scale": 1.1})

        chart_disp = wb.add_chart({"type": "line"})
        chart_disp.add_series(
            {
                "name": "Disp Objetivo (%)",
                "categories": ["Comparacion_Mensual", 1, c_mes, last, c_mes],
                "values": ["Comparacion_Mensual", 1, c_disp_obj, last, c_disp_obj],
            }
        )
        chart_disp.add_series(
            {
                "name": "Disp Real (%)",
                "categories": ["Comparacion_Mensual", 1, c_mes, last, c_mes],
                "values": ["Comparacion_Mensual", 1, c_disp_real, last, c_disp_real],
            }
        )
        chart_disp.set_title({"name": "Disponibilidad mensual: Objetivo vs Real"})
        chart_disp.set_y_axis({"name": "%"})
        chart_disp.set_legend({"position": "bottom"})
        ws_cmp.insert_chart("AA22", chart_disp, {"x_scale": 1.15, "y_scale": 1.1})

        chart_gap = wb.add_chart({"type": "column"})
        chart_gap.add_series(
            {
                "name": "Gap UEBD (pp)",
                "categories": ["Comparacion_Mensual", 1, c_mes, last, c_mes],
                "values": ["Comparacion_Mensual", 1, c_uebd_gap, last, c_uebd_gap],
            }
        )
        chart_gap.set_title({"name": "Gap mensual UEBD (pp)"})
        chart_gap.set_y_axis({"name": "pp"})
        ws_cmp.insert_chart("AA42", chart_gap, {"x_scale": 1.15, "y_scale": 1.1})

    if top_uebd:
        last = len(top_uebd)
        c_code = cols_top_uebd.index("codigo")
        c_impact = cols_top_uebd.index("impacto_atribuido_pp")
        chart_top_uebd = wb.add_chart({"type": "bar"})
        chart_top_uebd.add_series(
            {
                "name": "Impacto UEBD (pp)",
                "categories": ["Top_UEBD_Codigos", 1, c_code, last, c_code],
                "values": ["Top_UEBD_Codigos", 1, c_impact, last, c_impact],
            }
        )
        chart_top_uebd.set_title({"name": "Top codigos impacto UEBD"})
        chart_top_uebd.set_x_axis({"name": "pp"})
        ws_top_uebd.insert_chart("L2", chart_top_uebd, {"x_scale": 1.2, "y_scale": 1.25})

    if top_disp:
        last = len(top_disp)
        c_code = cols_top_disp.index("codigo")
        c_impact = cols_top_disp.index("impacto_atribuido_pp")
        chart_top_disp = wb.add_chart({"type": "bar"})
        chart_top_disp.add_series(
            {
                "name": "Impacto Disp (pp)",
                "categories": ["Top_Disp_Codigos", 1, c_code, last, c_code],
                "values": ["Top_Disp_Codigos", 1, c_impact, last, c_impact],
            }
        )
        chart_top_disp.set_title({"name": "Top codigos impacto Disponibilidad"})
        chart_top_disp.set_x_axis({"name": "pp"})
        ws_top_disp.insert_chart("L2", chart_top_disp, {"x_scale": 1.2, "y_scale": 1.25})

    if perdidas_resumen_rows:
        last = len(perdidas_resumen_rows)
        c_cat = cols_perd.index("concepto")
        c_horas = cols_perd.index("horas_perdidas")
        chart_loss = wb.add_chart({"type": "column"})
        chart_loss.add_series(
            {
                "name": "Horas perdidas",
                "categories": ["Perdidas_Resumen", 1, c_cat, last, c_cat],
                "values": ["Perdidas_Resumen", 1, c_horas, last, c_horas],
            }
        )
        chart_loss.set_title({"name": "Perdidas reales por componente"})
        chart_loss.set_y_axis({"name": "Horas"})
        ws_perd.insert_chart("H2", chart_loss, {"x_scale": 1.2, "y_scale": 1.1})

    # Mini dashboard en resumen
    if resumen_rows:
        row0 = resumen_rows[0]
        ws_resumen.write(3, 30, "UEBD Objetivo (%)", header_fmt)
        ws_resumen.write(4, 30, "UEBD Real (%)", header_fmt)
        ws_resumen.write(5, 30, "Gap UEBD (pp)", header_fmt)
        ws_resumen.write_number(3, 31, float(row0.get("uebd_objetivo_pct", 0.0)), pct_fmt)
        ws_resumen.write_number(4, 31, float(row0.get("uebd_real_pct", 0.0)), pct_fmt)
        ws_resumen.write_number(5, 31, float(row0.get("uebd_gap_pp", 0.0)), pct_fmt)

    wb.close()
    return True, ""


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Brecha 2025 vs Feb-2026 con atribucion por codigo.")
    parser.add_argument("input_2025", nargs="?", default=DEFAULT_2025_FILE, type=Path)
    parser.add_argument("input_2026", nargs="?", default=DEFAULT_2026_FILE, type=Path)
    parser.add_argument("--compare-year", type=int, default=2026)
    parser.add_argument("--compare-month", type=int, default=2)
    parser.add_argument("--uebd-objetivo", type=float, default=None, help="Objetivo UEBD (ej: 58 o 0.58).")
    parser.add_argument(
        "--disp-objetivo",
        type=float,
        default=None,
        help="Objetivo disponibilidad (ej: 82 o 0.82).",
    )
    parser.add_argument(
        "--mensual-2026",
        type=Path,
        default=None,
        help="Archivo mensual 2026 (CSV/XLSX) para leer objetivos de febrero.",
    )
    parser.add_argument(
        "--mensual-2026-sheet",
        default=None,
        help="Hoja del archivo mensual 2026 (si aplica).",
    )
    parser.add_argument(
        "--mensual-rig",
        default="TOTAL",
        help="Fila/rig a usar al leer objetivos mensuales (ej: TOTAL/FLOTA).",
    )
    parser.add_argument(
        "--mensual-a1-2026",
        type=Path,
        default=None,
        help=(
            "Mensual 2026 en formato A1 (Equipo/Indices/Unidad/Meses), "
            "ej: 'MENSUAL 2026' o 'MENSUAL 2026.xlsx'."
        ),
    )
    parser.add_argument(
        "--mensual-a1-sheet",
        default=None,
        help="Hoja del mensual A1 2026 (si aplica).",
    )
    parser.add_argument(
        "--mensual-a1-exclude-equipos",
        default="PF03,PFAR,PARR",
        help="Equipos a excluir al recalcular totales del mensual A1.",
    )
    parser.add_argument(
        "--horas-perdida-malla",
        type=float,
        default=0.0,
        help="Horas perdidas por disminucion de malla (manual, cuando se disponga).",
    )
    parser.add_argument("--top-codigos-grafico", type=int, default=12)
    parser.add_argument("--sin-graficos", action="store_true")
    parser.add_argument("--sin-excel", action="store_true")
    parser.add_argument("--output-dir", type=Path, default=Path("salidas_brecha_2025_vs_feb2026"))
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    path_2025 = resolve_input_csv_path(args.input_2025)
    path_2026 = resolve_input_csv_path(args.input_2026)

    base = aggregate_daily_data(path_2025, filter_year=2025, filter_month=None)
    comp_year = aggregate_daily_data(path_2026, filter_year=args.compare_year, filter_month=None)
    comp = aggregate_daily_data(path_2026, filter_year=args.compare_year, filter_month=args.compare_month)
    monthly_real_rows = aggregate_monthly_from_daily_rows(comp_year["daily_rows"], args.compare_year)

    uebd_obj = parse_ratio_value(args.uebd_objetivo)
    disp_obj = parse_ratio_value(args.disp_objetivo)
    source_uebd = "argumento"
    source_disp = "argumento"
    source_util = ""
    util_obj = None
    monthly_target_rows: List[Dict[str, Any]] = []

    if args.mensual_2026 is not None:
        mensual_path = args.mensual_2026
        if not mensual_path.exists() and mensual_path.suffix == "":
            monthly_candidates = [mensual_path.with_suffix(".xlsx"), mensual_path.with_suffix(".csv")]
            for c in monthly_candidates:
                if c.exists():
                    mensual_path = c
                    break
        if mensual_path.exists():
            monthly_target_rows = build_monthly_targets_from_long_table(
                mensual_path=mensual_path,
                year=args.compare_year,
                sheet_name=args.mensual_2026_sheet,
                rig_preference=args.mensual_rig,
            )
            extracted = extract_monthly_targets(
                mensual_path=mensual_path,
                year=args.compare_year,
                month=args.compare_month,
                sheet_name=args.mensual_2026_sheet,
                rig_preference=args.mensual_rig,
            )
            if uebd_obj is None and extracted["uebd_ratio"] is not None:
                uebd_obj = extracted["uebd_ratio"]
                source_uebd = f"mensual_2026:{mensual_path.name}"
            if disp_obj is None and extracted["disp_ratio"] is not None:
                disp_obj = extracted["disp_ratio"]
                source_disp = f"mensual_2026:{mensual_path.name}"

    if args.mensual_a1_2026 is not None:
        try:
            extracted_a1 = extract_targets_from_mensual_a1(
                mensual_a1_path=args.mensual_a1_2026,
                compare_year=args.compare_year,
                compare_month=args.compare_month,
                sheet_name=args.mensual_a1_sheet,
                excluded_equipos_csv=args.mensual_a1_exclude_equipos,
            )
            if extracted_a1.get("rows"):
                monthly_target_rows = extracted_a1["rows"]
            util_obj = extracted_a1.get("util_ratio")
            if util_obj is not None:
                source_util = f"mensual_a1:{Path(extracted_a1['source_path']).name}"
            if uebd_obj is None and extracted_a1["uebd_ratio"] is not None:
                uebd_obj = extracted_a1["uebd_ratio"]
                source_uebd = f"mensual_a1:{Path(extracted_a1['source_path']).name}"
            if disp_obj is None and extracted_a1["disp_ratio"] is not None:
                disp_obj = extracted_a1["disp_ratio"]
                source_disp = f"mensual_a1:{Path(extracted_a1['source_path']).name}"
        except Exception as exc:
            print(f"Advertencia: no fue posible leer mensual A1 2026 ({exc})")

    if uebd_obj is None:
        uebd_obj = base["uebd_ratio"]
        source_uebd = "fallback_baseline_2025"
    if disp_obj is None:
        disp_obj = base["disp_ratio"]
        source_disp = "fallback_baseline_2025"

    uebd_real = comp["uebd_ratio"]
    disp_real = comp["disp_ratio"]
    uebd_gap_pp = max((uebd_obj - uebd_real) * 100.0, 0.0)
    disp_gap_pp = max((disp_obj - disp_real) * 100.0, 0.0)

    compared_days = max(comp["day_count"], 1)
    avg_oper_hpd = comp["avg_hpd"]["horas_operativas"]
    avg_total_hpd = comp["avg_hpd"]["horas_totales"]

    uebd_delta_rows = build_code_delta_rows(base["uebd_code_hpd"], comp["uebd_code_hpd"])
    uebd_attr_rows = attribute_gap_to_codes(uebd_delta_rows, avg_oper_hpd, uebd_gap_pp, compared_days)

    disp_delta_rows = build_code_delta_rows(base["disp_code_hpd"], comp["disp_code_hpd"])
    disp_attr_rows = attribute_gap_to_codes(disp_delta_rows, avg_total_hpd, disp_gap_pp, compared_days)

    f09_rows = build_code_delta_rows(base["f09_code_hpd"], comp["f09_code_hpd"])
    for row in f09_rows:
        lost_hpd = row["delta_horas_dia_positivo"]
        row["impacto_potencial_uebd_pp"] = (lost_hpd / avg_oper_hpd * 100.0) if avg_oper_hpd > 0 else 0.0
        row["perdida_horas_mes"] = lost_hpd * compared_days

    perdida_horas_disponibilidad = disp_gap_pp / 100.0 * avg_total_hpd * compared_days
    perdida_horas_uebd = uebd_gap_pp / 100.0 * avg_oper_hpd * compared_days
    perdida_horas_rendimiento_f09 = sum(float(r["perdida_horas_mes"]) for r in f09_rows if r["delta_horas_dia_positivo"] > 0)
    perdida_horas_malla = max(float(args.horas_perdida_malla), 0.0)
    monthly_comparison_rows = build_monthly_comparison_rows(monthly_real_rows, monthly_target_rows)

    perdidas_resumen_rows = [
        {"concepto": "Perdida por disponibilidad", "horas_perdidas": perdida_horas_disponibilidad},
        {"concepto": "Perdida por UEBD", "horas_perdidas": perdida_horas_uebd},
        {"concepto": "Perdida por rendimiento F09", "horas_perdidas": perdida_horas_rendimiento_f09},
        {"concepto": "Perdida por malla", "horas_perdidas": perdida_horas_malla},
    ]

    resumen = [
        {
            "baseline_2025_archivo": path_2025.name,
            "comparado_archivo": path_2026.name,
            "comparado_periodo": f"{args.compare_year}-{args.compare_month:02d}",
            "dias_baseline_2025": base["day_count"],
            "dias_comparado": comp["day_count"],
            "uebd_objetivo_ratio": uebd_obj,
            "uebd_objetivo_pct": uebd_obj * 100.0,
            "uebd_real_ratio": uebd_real,
            "uebd_real_pct": uebd_real * 100.0,
            "uebd_gap_pp": uebd_gap_pp,
            "disponibilidad_objetivo_ratio": disp_obj,
            "disponibilidad_objetivo_pct": disp_obj * 100.0,
            "disponibilidad_real_ratio": disp_real,
            "disponibilidad_real_pct": disp_real * 100.0,
            "disponibilidad_gap_pp": disp_gap_pp,
            "horas_operativas_promedio_dia_comparado": avg_oper_hpd,
            "horas_totales_promedio_dia_comparado": avg_total_hpd,
            "perdida_horas_disponibilidad": perdida_horas_disponibilidad,
            "perdida_horas_uebd": perdida_horas_uebd,
            "perdida_horas_rendimiento_f09": perdida_horas_rendimiento_f09,
            "perdida_horas_malla": perdida_horas_malla,
            "perdida_horas_total": (
                perdida_horas_disponibilidad
                + perdida_horas_uebd
                + perdida_horas_rendimiento_f09
                + perdida_horas_malla
            ),
            "fuente_objetivo_uebd": source_uebd,
            "fuente_objetivo_disponibilidad": source_disp,
            "utilizacion_objetivo_ratio": util_obj if util_obj is not None else "",
            "utilizacion_objetivo_pct": (util_obj * 100.0) if util_obj is not None else "",
            "fuente_objetivo_utilizacion": source_util,
            "meses_con_comparacion_mensual": len(monthly_comparison_rows),
            "suma_impactos_uebd_pp": sum_field(uebd_attr_rows, "impacto_atribuido_pp"),
            "suma_impactos_disponibilidad_pp": sum_field(disp_attr_rows, "impacto_atribuido_pp"),
        }
    ]

    output_dir: Path = args.output_dir
    output_dir.mkdir(parents=True, exist_ok=True)

    write_csv(
        output_dir / "resumen_brecha_2025_vs_feb2026.csv",
        resumen,
        list(resumen[0].keys()),
    )
    write_csv(
        output_dir / "aporte_gap_uebd_por_codigo.csv",
        uebd_attr_rows,
        [
            "ranking_impacto",
            "codigo",
            "baseline_horas_dia",
            "comparado_horas_dia",
            "delta_horas_dia",
            "delta_horas_dia_positivo",
            "impacto_raw_pp",
            "factor_escalamiento",
            "impacto_atribuido_pp",
            "participacion_gap_pct",
            "impacto_acumulado_pp",
            "perdida_horas_dia_atribuida",
            "perdida_horas_mes_atribuida",
        ],
    )
    write_csv(
        output_dir / "aporte_gap_disponibilidad_por_codigo.csv",
        disp_attr_rows,
        [
            "ranking_impacto",
            "codigo",
            "baseline_horas_dia",
            "comparado_horas_dia",
            "delta_horas_dia",
            "delta_horas_dia_positivo",
            "impacto_raw_pp",
            "factor_escalamiento",
            "impacto_atribuido_pp",
            "participacion_gap_pct",
            "impacto_acumulado_pp",
            "perdida_horas_dia_atribuida",
            "perdida_horas_mes_atribuida",
        ],
    )
    write_csv(
        output_dir / "top_aporte_gap_uebd_por_codigo.csv",
        [r for r in uebd_attr_rows if float(r.get("impacto_atribuido_pp", 0.0)) > 0][:20],
        [
            "ranking_impacto",
            "codigo",
            "impacto_atribuido_pp",
            "participacion_gap_pct",
            "impacto_acumulado_pp",
            "perdida_horas_mes_atribuida",
            "delta_horas_dia_positivo",
        ],
    )
    write_csv(
        output_dir / "top_aporte_gap_disponibilidad_por_codigo.csv",
        [r for r in disp_attr_rows if float(r.get("impacto_atribuido_pp", 0.0)) > 0][:20],
        [
            "ranking_impacto",
            "codigo",
            "impacto_atribuido_pp",
            "participacion_gap_pct",
            "impacto_acumulado_pp",
            "perdida_horas_mes_atribuida",
            "delta_horas_dia_positivo",
        ],
    )
    write_csv(
        output_dir / "perdida_rendimiento_f09_por_codigo.csv",
        f09_rows,
        [
            "codigo",
            "baseline_horas_dia",
            "comparado_horas_dia",
            "delta_horas_dia",
            "delta_horas_dia_positivo",
            "impacto_potencial_uebd_pp",
            "perdida_horas_mes",
        ],
    )
    write_csv(
        output_dir / "diario_baseline_2025.csv",
        base["daily_rows"],
        [
            "fecha_operativa",
            "horas_totales",
            "horas_operativas",
            "horas_efectivo",
            "horas_reserva",
            "horas_mant_programada",
            "horas_mant_no_programada",
            "horas_otras",
            "uebd_ratio",
            "uebd_pct",
            "disponibilidad_ratio",
            "disponibilidad_pct",
        ],
    )
    write_csv(
        output_dir / "comparacion_mensual_objetivo_vs_real.csv",
        monthly_comparison_rows,
        [
            "anio",
            "mes_num",
            "mes",
            "dias_reales_con_datos",
            "horas_totales_real",
            "horas_operativas_real",
            "horas_efectivo_real",
            "disponibilidad_obj_ratio",
            "disponibilidad_obj_pct",
            "disponibilidad_real_ratio",
            "disponibilidad_real_pct",
            "disponibilidad_gap_pp",
            "utilizacion_obj_ratio",
            "utilizacion_obj_pct",
            "utilizacion_real_ratio",
            "utilizacion_real_pct",
            "utilizacion_gap_pp",
            "uebd_obj_ratio",
            "uebd_obj_pct",
            "uebd_real_ratio",
            "uebd_real_pct",
            "uebd_gap_pp",
            "perdida_horas_uebd_mes",
            "perdida_horas_disponibilidad_mes",
        ],
    )
    write_csv(
        output_dir / "resumen_perdidas_reales_componentes.csv",
        perdidas_resumen_rows,
        ["concepto", "horas_perdidas"],
    )
    write_csv(
        output_dir / "diario_comparado_feb2026.csv",
        comp["daily_rows"],
        [
            "fecha_operativa",
            "horas_totales",
            "horas_operativas",
            "horas_efectivo",
            "horas_reserva",
            "horas_mant_programada",
            "horas_mant_no_programada",
            "horas_otras",
            "uebd_ratio",
            "uebd_pct",
            "disponibilidad_ratio",
            "disponibilidad_pct",
        ],
    )

    chart_messages: List[str] = []
    if not args.sin_graficos:
        uebd_contrib = top_contributions(uebd_attr_rows, top_n=args.top_codigos_grafico)
        disp_contrib = top_contributions(disp_attr_rows, top_n=args.top_codigos_grafico)
        ok_u, err_u = draw_gap_waterfall(
            output_path=output_dir / "graficos" / "cascada_brecha_uebd_codigos.png",
            title="Brecha UEBD: Objetivo vs Real (atribuida por codigo)",
            start_pct=uebd_obj * 100.0,
            final_pct=uebd_real * 100.0,
            contributions=uebd_contrib,
        )
        ok_d, err_d = draw_gap_waterfall(
            output_path=output_dir / "graficos" / "cascada_brecha_disponibilidad_codigos.png",
            title="Brecha Disponibilidad: Objetivo vs Real (atribuida por codigo)",
            start_pct=disp_obj * 100.0,
            final_pct=disp_real * 100.0,
            contributions=disp_contrib,
        )
        chart_messages.append(
            "Cascada UEBD: generado" if ok_u else f"Cascada UEBD: omitido ({err_u})"
        )
        chart_messages.append(
            "Cascada Disponibilidad: generado" if ok_d else f"Cascada Disponibilidad: omitido ({err_d})"
        )
        ok_line_u, err_line_u = draw_monthly_obj_vs_real_lines(
            output_path=output_dir / "graficos" / "mensual_uebd_objetivo_vs_real.png",
            rows=monthly_comparison_rows,
            obj_key_pct="uebd_obj_pct",
            real_key_pct="uebd_real_pct",
            title="UEBD mensual: objetivo vs real",
        )
        chart_messages.append(
            "Grafico mensual UEBD objetivo vs real: generado"
            if ok_line_u
            else f"Grafico mensual UEBD objetivo vs real: omitido ({err_line_u})"
        )
        ok_line_d, err_line_d = draw_monthly_obj_vs_real_lines(
            output_path=output_dir / "graficos" / "mensual_disp_objetivo_vs_real.png",
            rows=monthly_comparison_rows,
            obj_key_pct="disponibilidad_obj_pct",
            real_key_pct="disponibilidad_real_pct",
            title="Disponibilidad mensual: objetivo vs real",
        )
        chart_messages.append(
            "Grafico mensual disponibilidad objetivo vs real: generado"
            if ok_line_d
            else f"Grafico mensual disponibilidad objetivo vs real: omitido ({err_line_d})"
        )
        ok_gap, err_gap = draw_monthly_gap_bars(
            output_path=output_dir / "graficos" / "mensual_gap_uebd_pp.png",
            rows=monthly_comparison_rows,
            gap_key_pp="uebd_gap_pp",
            title="Gap mensual UEBD (pp)",
        )
        chart_messages.append(
            "Grafico mensual gap UEBD: generado"
            if ok_gap
            else f"Grafico mensual gap UEBD: omitido ({err_gap})"
        )
        ok_top_u, err_top_u = draw_top_code_impact_bars(
            output_path=output_dir / "graficos" / "top_codigos_impacto_uebd.png",
            rows=uebd_attr_rows,
            title="Top codigos que explican perdida UEBD (pp)",
            top_n=max(args.top_codigos_grafico, 10),
        )
        chart_messages.append(
            "Grafico top codigos UEBD: generado"
            if ok_top_u
            else f"Grafico top codigos UEBD: omitido ({err_top_u})"
        )
        ok_top_d, err_top_d = draw_top_code_impact_bars(
            output_path=output_dir / "graficos" / "top_codigos_impacto_disponibilidad.png",
            rows=disp_attr_rows,
            title="Top codigos que explican perdida Disponibilidad (pp)",
            top_n=max(args.top_codigos_grafico, 10),
        )
        chart_messages.append(
            "Grafico top codigos disponibilidad: generado"
            if ok_top_d
            else f"Grafico top codigos disponibilidad: omitido ({err_top_d})"
        )
    else:
        chart_messages.append("Graficos deshabilitados por parametro.")

    excel_message = ""
    if not args.sin_excel:
        ok_x, err_x = write_excel(
            output_xlsx=output_dir / "reporte_brecha_2025_vs_feb2026.xlsx",
            resumen_rows=resumen,
            uebd_rows=uebd_attr_rows,
            disp_rows=disp_attr_rows,
            f09_rows=f09_rows,
            daily_2025_rows=base["daily_rows"],
            daily_feb_rows=comp["daily_rows"],
            monthly_comparison_rows=monthly_comparison_rows,
            perdidas_resumen_rows=perdidas_resumen_rows,
        )
        excel_message = (
            "Excel: generado (reporte_brecha_2025_vs_feb2026.xlsx)"
            if ok_x
            else f"Excel: omitido ({err_x})"
        )
    else:
        excel_message = "Excel deshabilitado por parametro."

    print("Analisis de brecha finalizado.")
    print(f"Baseline 2025: {path_2025}")
    print(f"Comparado: {path_2026} | periodo {args.compare_year}-{args.compare_month:02d}")
    print(f"Dias baseline 2025: {base['day_count']}")
    print(f"Dias comparado: {comp['day_count']}")
    print(f"UEBD objetivo: {uebd_obj * 100.0:.2f}% | real: {uebd_real * 100.0:.2f}% | gap: {uebd_gap_pp:.2f} pp")
    print(
        "Disponibilidad objetivo: "
        f"{disp_obj * 100.0:.2f}% | real: {disp_real * 100.0:.2f}% | gap: {disp_gap_pp:.2f} pp"
    )
    print(f"Suma impactos atribuidos UEBD: {sum_field(uebd_attr_rows, 'impacto_atribuido_pp'):.4f} pp")
    print(
        "Suma impactos atribuidos Disponibilidad: "
        f"{sum_field(disp_attr_rows, 'impacto_atribuido_pp'):.4f} pp"
    )
    print(f"Meses con comparacion mensual objetivo vs real: {len(monthly_comparison_rows)}")
    print(f"Perdida horas Disponibilidad: {perdida_horas_disponibilidad:.2f}")
    print(f"Perdida horas UEBD: {perdida_horas_uebd:.2f}")
    print(f"Perdida horas Rendimiento F09: {perdida_horas_rendimiento_f09:.2f}")
    print(f"Perdida horas Malla (manual): {perdida_horas_malla:.2f}")
    print(f"Salida: {output_dir}")
    for msg in chart_messages:
        print(msg)
    print(excel_message)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
