#!/usr/bin/env python3
"""
Comparador mensual 2025 vs 2026 por perforadora.

Lee tablas mensuales en formato A1:
  A: Equipo
  B: Indices
  C: Unidad
  D..: Meses (Enero, Febrero, ..., Diciembre)

Capacidades:
- Comparar valores mensuales por perforadora entre 2025 y 2026.
- Excluir equipos (por ejemplo PF03 y PFAR/PARR).
- Recalcular totales mensuales SIN los equipos excluidos.
- Comparar totales recalculados entre 2025 y 2026.
"""

from __future__ import annotations

import argparse
import csv
import re
import unicodedata
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Set, Tuple

MONTHS_ES = {
    "enero": 1,
    "febrero": 2,
    "marzo": 3,
    "abril": 4,
    "mayo": 5,
    "junio": 6,
    "julio": 7,
    "agosto": 8,
    "septiembre": 9,
    "setiembre": 9,
    "octubre": 10,
    "noviembre": 11,
    "diciembre": 12,
}

MONTH_NAME_BY_NUM = {
    1: "Enero",
    2: "Febrero",
    3: "Marzo",
    4: "Abril",
    5: "Mayo",
    6: "Junio",
    7: "Julio",
    8: "Agosto",
    9: "Septiembre",
    10: "Octubre",
    11: "Noviembre",
    12: "Diciembre",
}

DEFAULT_MENSUAL_2025 = "MENSUAL 2025 (1)"
DEFAULT_MENSUAL_2026 = "MENSUAL 2026"


def normalize_text(value: Any) -> str:
    txt = str(value or "").strip().lower()
    dec = unicodedata.normalize("NFKD", txt)
    return "".join(ch for ch in dec if not unicodedata.combining(ch))


def normalize_equipo(value: Any) -> str:
    txt = normalize_text(value).upper()
    # Mantiene solo alfanumericos para equiparar PF-03, PF03, pf03.
    return re.sub(r"[^A-Z0-9]", "", txt)


def canonical_index(indice_raw: Any) -> str:
    idx = normalize_text(indice_raw)
    if "dispon" in idx:
        return "disponibilidad"
    if "util" in idx:
        return "utilizacion"
    if "rend" in idx:
        return "rendimiento"
    if "metro" in idx:
        return "metros"
    if "hora" in idx and ("ef" in idx or "efect" in idx):
        return "horas_efectivas"
    return idx.replace(" ", "_")


def parse_numeric(value: Any) -> Optional[float]:
    if value is None:
        return None
    txt = str(value).strip()
    if not txt:
        return None
    txt = txt.replace("%", "").replace(",", "")
    try:
        return float(txt)
    except ValueError:
        return None


def parse_month_header(value: Any) -> Optional[int]:
    txt = normalize_text(value)
    if txt.isdigit():
        m = int(txt)
        if 1 <= m <= 12:
            return m
    return MONTHS_ES.get(txt)


def is_total_equipo(equipo_norm: str) -> bool:
    return equipo_norm in {"TOTAL", "FLOTA"}


def resolve_table_file(path: Path) -> Path:
    if path.exists():
        return path
    if path.suffix:
        raise FileNotFoundError(f"No existe archivo: {path}")

    candidates = [
        path.with_suffix(".xlsx"),
        path.with_suffix(".xlsm"),
        path.with_suffix(".xls"),
        path.with_suffix(".csv"),
    ]
    for c in candidates:
        if c.exists():
            return c
    attempted = ", ".join(str(c) for c in candidates)
    raise FileNotFoundError(f"No existe archivo: {path}. Intentados: {attempted}")


def read_table_from_csv(path: Path) -> List[Dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        sample = f.read(4096)
        f.seek(0)
        delimiter = ";" if sample.count(";") >= sample.count(",") else ","
        reader = csv.DictReader(f, delimiter=delimiter)
        return [dict(r) for r in reader]


def read_table_from_xlsx(path: Path, sheet_name: Optional[str]) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover
        raise RuntimeError("openpyxl no disponible para leer xlsx.") from exc

    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if (sheet_name and sheet_name in wb.sheetnames) else wb[wb.sheetnames[0]]

    # Formato requerido: encabezado en A1.
    header = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    # recorta columnas vacias del final
    while header and (header[-1] is None or str(header[-1]).strip() == ""):
        header.pop()
    headers = [str(h).strip() if h is not None else "" for h in header]

    rows: List[Dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        values = [ws.cell(row=r, column=c).value for c in range(1, len(headers) + 1)]
        if not any(v is not None and str(v).strip() for v in values):
            continue
        row: Dict[str, Any] = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            row[h] = values[i] if i < len(values) else None
        rows.append(row)
    return rows


def read_table(path: Path, sheet_name: Optional[str]) -> List[Dict[str, Any]]:
    if path.suffix.lower() == ".csv":
        return read_table_from_csv(path)
    if path.suffix.lower() in {".xlsx", ".xlsm", ".xltx", ".xltm", ".xls"}:
        return read_table_from_xlsx(path, sheet_name)
    raise ValueError(f"Formato no soportado: {path.suffix}")


def detect_column(headers: Sequence[str], candidates: Sequence[str]) -> Optional[str]:
    nh = {h: normalize_text(h) for h in headers}
    for c in candidates:
        nc = normalize_text(c)
        for h, hv in nh.items():
            if nc in hv:
                return h
    return None


def parse_mensual_records(table_rows: List[Dict[str, Any]], source_year: int) -> List[Dict[str, Any]]:
    if not table_rows:
        return []
    headers = list(table_rows[0].keys())
    col_equipo = detect_column(headers, ["equipo", "rig", "perforadora"])
    col_indice = detect_column(headers, ["indices", "indice"])
    col_unidad = detect_column(headers, ["unidad"])
    if not col_equipo or not col_indice:
        raise ValueError("No se pudieron detectar columnas 'Equipo' e 'Indices' desde A1.")

    month_cols: List[Tuple[str, int]] = []
    for h in headers:
        if h in {col_equipo, col_indice, col_unidad}:
            continue
        m = parse_month_header(h)
        if m is not None:
            month_cols.append((h, m))
    if not month_cols:
        raise ValueError("No se detectaron columnas de meses (Enero..Diciembre) desde A1.")

    records: List[Dict[str, Any]] = []
    for row in table_rows:
        equipo_raw = row.get(col_equipo)
        indice_raw = row.get(col_indice)
        if equipo_raw is None or indice_raw is None:
            continue
        equipo_label = str(equipo_raw).strip()
        if not equipo_label:
            continue
        indice_label = str(indice_raw).strip()
        if not indice_label:
            continue
        unidad_label = str(row.get(col_unidad) or "").strip() if col_unidad else ""

        for month_col, month_num in month_cols:
            raw_val = row.get(month_col)
            val = parse_numeric(raw_val)
            if val is None:
                continue
            records.append(
                {
                    "source_year": source_year,
                    "equipo": equipo_label,
                    "equipo_norm": normalize_equipo(equipo_label),
                    "indice": indice_label,
                    "indice_key": canonical_index(indice_label),
                    "unidad": unidad_label,
                    "mes_num": month_num,
                    "mes": MONTH_NAME_BY_NUM.get(month_num, str(month_num)),
                    "valor": val,
                }
            )
    return records


def pct_to_ratio(val: Optional[float]) -> Optional[float]:
    if val is None:
        return None
    if val > 1:
        return val / 100.0
    return val


def compare_by_rig(
    records_2025: List[Dict[str, Any]],
    records_2026: List[Dict[str, Any]],
    excluded_equipo_norms: Set[str],
) -> List[Dict[str, Any]]:
    map_2025 = {}
    map_2026 = {}

    for r in records_2025:
        key = (r["equipo_norm"], r["equipo"], r["indice_key"], r["indice"], r["unidad"], r["mes_num"], r["mes"])
        map_2025[key] = float(r["valor"])
    for r in records_2026:
        key = (r["equipo_norm"], r["equipo"], r["indice_key"], r["indice"], r["unidad"], r["mes_num"], r["mes"])
        map_2026[key] = float(r["valor"])

    # Homologa por equipo_norm + indice + mes
    keys_norm = set()
    for k in map_2025:
        keys_norm.add((k[0], k[2], k[5]))
    for k in map_2026:
        keys_norm.add((k[0], k[2], k[5]))

    # etiquetas de salida por prioridad 2026 luego 2025
    label_by_norm: Dict[str, str] = {}
    indice_label: Dict[str, str] = {}
    unidad_by_indice: Dict[str, str] = {}
    mes_label: Dict[int, str] = {}
    for k in list(map_2026.keys()) + list(map_2025.keys()):
        label_by_norm[k[0]] = label_by_norm.get(k[0], k[1])
        indice_label[k[2]] = indice_label.get(k[2], k[3])
        unidad_by_indice[k[2]] = unidad_by_indice.get(k[2], k[4])
        mes_label[k[5]] = mes_label.get(k[5], k[6])

    out: List[Dict[str, Any]] = []
    for equipo_norm, idx_key, mes_num in sorted(keys_norm, key=lambda x: (x[2], x[0], x[1])):
        if is_total_equipo(equipo_norm):
            continue
        if equipo_norm in excluded_equipo_norms:
            continue

        equipo = label_by_norm.get(equipo_norm, equipo_norm)
        idx = indice_label.get(idx_key, idx_key)
        unidad = unidad_by_indice.get(idx_key, "")
        mes = mes_label.get(mes_num, MONTH_NAME_BY_NUM.get(mes_num, str(mes_num)))

        v25 = None
        v26 = None
        # busca con cualquier etiqueta equivalente
        for k, v in map_2025.items():
            if k[0] == equipo_norm and k[2] == idx_key and k[5] == mes_num:
                v25 = v
                break
        for k, v in map_2026.items():
            if k[0] == equipo_norm and k[2] == idx_key and k[5] == mes_num:
                v26 = v
                break

        if v25 is None and v26 is None:
            continue

        delta_abs = (v26 - v25) if (v25 is not None and v26 is not None) else None
        delta_rel_pct = ((v26 - v25) / v25 * 100.0) if (v25 not in {None, 0} and v26 is not None) else None
        out.append(
            {
                "mes_num": mes_num,
                "mes": mes,
                "equipo": equipo,
                "indice": idx,
                "unidad": unidad,
                "valor_2025": v25,
                "valor_2026": v26,
                "delta_abs": delta_abs,
                "delta_rel_pct": delta_rel_pct,
            }
        )
    return out


def recalculate_totals(
    records: List[Dict[str, Any]],
    excluded_equipo_norms: Set[str],
    source_year: int,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    # mapa base
    value_map: Dict[Tuple[str, str, int], float] = {}
    equipo_label: Dict[str, str] = {}
    indice_label: Dict[str, str] = {}
    unidad_by_indice: Dict[str, str] = {}

    for r in records:
        eq = r["equipo_norm"]
        ik = r["indice_key"]
        m = int(r["mes_num"])
        value_map[(eq, ik, m)] = float(r["valor"])
        equipo_label[eq] = r["equipo"]
        indice_label[ik] = r["indice"]
        if r["unidad"]:
            unidad_by_indice[ik] = r["unidad"]

    months = sorted({int(r["mes_num"]) for r in records})
    equipos = sorted({r["equipo_norm"] for r in records if not is_total_equipo(r["equipo_norm"])})

    recalculated: List[Dict[str, Any]] = []
    comparison: List[Dict[str, Any]] = []

    for m in months:
        horas_ef_sum = 0.0
        metros_sum = 0.0
        oper_sum = 0.0
        total_sum = 0.0
        included_count = 0

        for eq in equipos:
            if eq in excluded_equipo_norms:
                continue
            he = value_map.get((eq, "horas_efectivas", m))
            me = value_map.get((eq, "metros", m))
            ut = value_map.get((eq, "utilizacion", m))
            dp = value_map.get((eq, "disponibilidad", m))

            if he is None:
                continue

            included_count += 1
            horas_ef_sum += he
            if me is not None:
                metros_sum += me

            ut_ratio = pct_to_ratio(ut) if ut is not None else None
            if ut_ratio is not None and ut_ratio > 0:
                oper = he / ut_ratio
                oper_sum += oper
                dp_ratio = pct_to_ratio(dp) if dp is not None else None
                if dp_ratio is not None and dp_ratio > 0:
                    total_sum += oper / dp_ratio

        util_total_pct = (horas_ef_sum / oper_sum * 100.0) if oper_sum > 0 else None
        disp_total_pct = (oper_sum / total_sum * 100.0) if total_sum > 0 else None
        rend_total = (metros_sum / horas_ef_sum) if horas_ef_sum > 0 else None

        month_name = MONTH_NAME_BY_NUM.get(m, str(m))
        rec_rows = [
            ("disponibilidad", "Disponibilidad", "%", disp_total_pct),
            ("utilizacion", "Utilización", "%", util_total_pct),
            ("rendimiento", "Rendimiento", "m/hr efect", rend_total),
            ("metros", "Metros", "m", metros_sum),
            ("horas_efectivas", "Horas Efectivas", "Hr ef", horas_ef_sum),
        ]
        for ik, idx_label_default, unit_default, val in rec_rows:
            recalculated.append(
                {
                    "anio": source_year,
                    "mes_num": m,
                    "mes": month_name,
                    "indice_key": ik,
                    "indice": indice_label.get(ik, idx_label_default),
                    "unidad": unidad_by_indice.get(ik, unit_default),
                    "valor_total_recalculado_sin_excluidas": val,
                    "equipos_incluidos": included_count,
                    "equipos_excluidos": len(excluded_equipo_norms),
                }
            )

            total_original = None
            for total_key in ("TOTAL", "FLOTA"):
                if (total_key, ik, m) in value_map:
                    total_original = value_map[(total_key, ik, m)]
                    break
            comparison.append(
                {
                    "anio": source_year,
                    "mes_num": m,
                    "mes": month_name,
                    "indice_key": ik,
                    "indice": indice_label.get(ik, idx_label_default),
                    "unidad": unidad_by_indice.get(ik, unit_default),
                    "total_original_tabla": total_original,
                    "total_recalculado_sin_excluidas": val,
                    "diferencia_recalculado_menos_original": (
                        (val - total_original) if (val is not None and total_original is not None) else None
                    ),
                }
            )

    return recalculated, comparison


def compare_recalculated_totals(
    rec_2025: List[Dict[str, Any]],
    rec_2026: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    map25 = {(r["mes_num"], r["indice_key"]): r for r in rec_2025}
    map26 = {(r["mes_num"], r["indice_key"]): r for r in rec_2026}
    keys = sorted(set(map25) | set(map26), key=lambda x: (x[0], x[1]))
    rows: List[Dict[str, Any]] = []
    for key in keys:
        r25 = map25.get(key)
        r26 = map26.get(key)
        mes_num, idx_key = key
        mes = MONTH_NAME_BY_NUM.get(mes_num, str(mes_num))
        idx = (r26 or r25 or {}).get("indice", idx_key)
        unidad = (r26 or r25 or {}).get("unidad", "")
        v25 = r25.get("valor_total_recalculado_sin_excluidas") if r25 else None
        v26 = r26.get("valor_total_recalculado_sin_excluidas") if r26 else None
        delta = (v26 - v25) if (v25 is not None and v26 is not None) else None
        delta_rel = ((v26 - v25) / v25 * 100.0) if (v25 not in {None, 0} and v26 is not None) else None
        rows.append(
            {
                "mes_num": mes_num,
                "mes": mes,
                "indice_key": idx_key,
                "indice": idx,
                "unidad": unidad,
                "total_2025_sin_excluidas": v25,
                "total_2026_sin_excluidas": v26,
                "delta_abs": delta,
                "delta_rel_pct": delta_rel,
            }
        )
    return rows


def format_value(value: Any) -> str:
    if isinstance(value, float):
        return f"{value:.6f}"
    return str(value)


def write_csv(path: Path, rows: List[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        with path.open("w", encoding="utf-8", newline="") as f:
            f.write("sin_datos\n")
        return
    fields = list(rows[0].keys())
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields, delimiter=";")
        writer.writeheader()
        for r in rows:
            writer.writerow({k: format_value(r.get(k, "")) for k in fields})


def try_import_xlsxwriter():
    try:
        import xlsxwriter

        return xlsxwriter, ""
    except Exception as exc:  # pragma: no cover
        return None, str(exc)


def write_excel_report(
    output_path: Path,
    by_rig_rows: List[Dict[str, Any]],
    rec25_rows: List[Dict[str, Any]],
    rec26_rows: List[Dict[str, Any]],
    cmp25_rows: List[Dict[str, Any]],
    cmp26_rows: List[Dict[str, Any]],
    rec_cmp_rows: List[Dict[str, Any]],
) -> Tuple[bool, str]:
    xlsxwriter, err = try_import_xlsxwriter()
    if xlsxwriter is None:
        return False, err

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = xlsxwriter.Workbook(str(output_path))
    hfmt = wb.add_format({"bold": True, "bg_color": "#D9E1F2", "border": 1})
    nfmt = wb.add_format({"num_format": "0.0000"})
    pfmt = wb.add_format({"num_format": "0.00"})

    def add_sheet(name: str, rows: List[Dict[str, Any]]):
        ws = wb.add_worksheet(name[:31])
        if not rows:
            ws.write(0, 0, "Sin datos")
            return ws
        cols = list(rows[0].keys())
        for c, col in enumerate(cols):
            ws.write(0, c, col, hfmt)
        for r, row in enumerate(rows, start=1):
            for c, col in enumerate(cols):
                val = row.get(col, "")
                if isinstance(val, float):
                    fmt = pfmt if col.endswith("_pct") else nfmt
                    ws.write_number(r, c, val, fmt)
                elif isinstance(val, int):
                    ws.write_number(r, c, val)
                else:
                    ws.write(r, c, str(val))
        for c, col in enumerate(cols):
            ws.set_column(c, c, max(12, min(38, len(col) + 2)))
        return ws

    add_sheet("Comparado_Rig", by_rig_rows)
    add_sheet("Totales_Recalc_2025", rec25_rows)
    add_sheet("Totales_Recalc_2026", rec26_rows)
    add_sheet("Vs_Original_2025", cmp25_rows)
    add_sheet("Vs_Original_2026", cmp26_rows)
    ws_cmp = add_sheet("Totales_2025_vs_2026", rec_cmp_rows)

    # grafico simple por mes de Horas Efectivas total recalculado
    if rec_cmp_rows:
        he_rows = [i for i, r in enumerate(rec_cmp_rows, start=1) if r.get("indice_key") == "horas_efectivas"]
        if he_rows:
            cols = list(rec_cmp_rows[0].keys())
            c_mes = cols.index("mes")
            c_25 = cols.index("total_2025_sin_excluidas")
            c_26 = cols.index("total_2026_sin_excluidas")
            chart = wb.add_chart({"type": "line"})
            chart.add_series(
                {
                    "name": "2025 Horas Efectivas",
                    "categories": ["Totales_2025_vs_2026", he_rows[0], c_mes, he_rows[-1], c_mes],
                    "values": ["Totales_2025_vs_2026", he_rows[0], c_25, he_rows[-1], c_25],
                }
            )
            chart.add_series(
                {
                    "name": "2026 Horas Efectivas",
                    "categories": ["Totales_2025_vs_2026", he_rows[0], c_mes, he_rows[-1], c_mes],
                    "values": ["Totales_2025_vs_2026", he_rows[0], c_26, he_rows[-1], c_26],
                }
            )
            chart.set_title({"name": "Horas Efectivas (totales recalculados)"})
            chart.set_legend({"position": "bottom"})
            ws_cmp.insert_chart("K2", chart, {"x_scale": 1.2, "y_scale": 1.2})

    wb.close()
    return True, ""


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Compara mensuales 2025 vs 2026 por perforadora y recalcula totales excluyendo equipos."
    )
    parser.add_argument("mensual_2025", nargs="?", type=Path, default=Path(DEFAULT_MENSUAL_2025))
    parser.add_argument("mensual_2026", nargs="?", type=Path, default=Path(DEFAULT_MENSUAL_2026))
    parser.add_argument("--sheet-2025", default=None)
    parser.add_argument("--sheet-2026", default=None)
    parser.add_argument(
        "--exclude-equipos",
        default="PF03,PFAR,PARR",
        help="Lista separada por coma de equipos a excluir de los totales.",
    )
    parser.add_argument("--output-dir", type=Path, default=Path("salidas_comparador_mensual"))
    parser.add_argument("--sin-excel", action="store_true")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    mensual_2025_path = resolve_table_file(args.mensual_2025)
    mensual_2026_path = resolve_table_file(args.mensual_2026)

    excluidos = {normalize_equipo(x) for x in args.exclude_equipos.split(",") if x.strip()}

    rows_2025 = parse_mensual_records(read_table(mensual_2025_path, args.sheet_2025), source_year=2025)
    rows_2026 = parse_mensual_records(read_table(mensual_2026_path, args.sheet_2026), source_year=2026)

    by_rig_cmp = compare_by_rig(rows_2025, rows_2026, excluded_equipo_norms=excluidos)
    rec_2025, cmp_2025_vs_orig = recalculate_totals(rows_2025, excluded_equipo_norms=excluidos, source_year=2025)
    rec_2026, cmp_2026_vs_orig = recalculate_totals(rows_2026, excluded_equipo_norms=excluidos, source_year=2026)
    rec_cmp = compare_recalculated_totals(rec_2025, rec_2026)

    out = args.output_dir
    write_csv(out / "mensual_comparado_por_perforadora.csv", by_rig_cmp)
    write_csv(out / "totales_recalculados_sin_excluidas_2025.csv", rec_2025)
    write_csv(out / "totales_recalculados_sin_excluidas_2026.csv", rec_2026)
    write_csv(out / "totales_2025_vs_original_2025.csv", cmp_2025_vs_orig)
    write_csv(out / "totales_2026_vs_original_2026.csv", cmp_2026_vs_orig)
    write_csv(out / "totales_recalculados_2025_vs_2026.csv", rec_cmp)

    excel_msg = ""
    if not args.sin_excel:
        ok, err = write_excel_report(
            output_path=out / "reporte_mensual_comparado_sin_excluidas.xlsx",
            by_rig_rows=by_rig_cmp,
            rec25_rows=rec_2025,
            rec26_rows=rec_2026,
            cmp25_rows=cmp_2025_vs_orig,
            cmp26_rows=cmp_2026_vs_orig,
            rec_cmp_rows=rec_cmp,
        )
        excel_msg = (
            "Excel: generado (reporte_mensual_comparado_sin_excluidas.xlsx)"
            if ok
            else f"Excel: omitido ({err})"
        )
    else:
        excel_msg = "Excel deshabilitado por parametro."

    print("Comparador mensual finalizado.")
    print(f"Mensual 2025: {mensual_2025_path}")
    print(f"Mensual 2026: {mensual_2026_path}")
    print(f"Equipos excluidos: {', '.join(sorted(excluidos)) if excluidos else '(ninguno)'}")
    print(f"Filas comparadas por perforadora: {len(by_rig_cmp)}")
    print(f"Filas totales recalculados 2025: {len(rec_2025)}")
    print(f"Filas totales recalculados 2026: {len(rec_2026)}")
    print(f"Filas comparacion totales 2025 vs 2026: {len(rec_cmp)}")
    print(f"Salida: {out}")
    print(excel_msg)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
